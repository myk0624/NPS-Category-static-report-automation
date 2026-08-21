import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from html import escape
import traceback
import csv
import time
from openpyxl.styles import Font, PatternFill, Border


# ──────────────────────────────────────────────────────────────────────────────
# Column index helper
# ──────────────────────────────────────────────────────────────────────────────

def ci(col: str) -> int:
    """Excel column letter(s) → 0-based int.  A=0, B=1, AA=26, BL=63 …"""
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


# ──────────────────────────────────────────────────────────────────────────────
# CSV 파일 종류 판별 (파일명이 아닌 컬럼 구조 기반)
# ──────────────────────────────────────────────────────────────────────────────

MEDIA_COLUMNS = [
    'Date', 'Year', 'Month', 'Week', 'Media', 'Campaign', 'Ad Group', 'ad',
    'Spend (gross)', 'Impressions', 'Click', 'Video Views', 'Reach',
    'AF Installs', 'AF Reinstalls', 'AF Revenue', 'AF Order Amount',
    'AF add_to_cart', 'AF Coupon', 'pb_order_now_delivery_quick',
    'pb_order_now_delivery_mart', 'pb_order_now_delivery_all',
    'pb_order_kurlynmart', 'pb_order_kurlynmart(Unique)',
    'pb_order_kurlynmart_revenue', 'pb_view_product_kurlynmart',
    'pb_view_product_kurlynmart(Unique)', 'kurly_view_home',
    'kurly_view_home(Unique)', 'first_order', 'temp_promo1', 'temp_promo2',
    'temp_event1', 'temp_event1(Unique)', 'OS', 'Campaign Theme',
    'Campaign Objective', 'Ad Product', 'Campaign Details', 'Targeting',
    'Gender', 'Age', 'Detailed Targeting', 'Creative Live Date',
    'Creative Format', 'Creative Type', 'Dimension', 'USP Category',
    'USP Brand', 'USP', '애드코드', 'Creative Full Name', '대구분',
    're-engagement', 'gross(net)', 'Campaign Theme(ADEF)',
    'Campaign Objective(ADEF)', 'Ad Product(ADEF)',
    'Detailed Targeting(ADEF)', 'USP Category(ADEF)', 'USP Brand(ADEF)',
    'USP(ADEF)', 'AF Order Count(Adef)', '집약형(Adef)', 'Install(SKAN)',
    '프로모션', '소재이미지', '소재카피', '소구점',
]

CATEGORY_EXTRA_COLUMNS = [
    '카테고리 구매 unique', '카테고리 구매 quantity', '카테고리 구매 price',
    '패션 unique', '패션 quantity', '패션 price',
    '뷰티 unique', '뷰티 quantity', '뷰티 price',
    '디지털가전 unique', '디지털가전 quantity', '디지털가전 price',
    '가구 unique', '가구 quantity', '가구 price',
    '키즈 unique', '키즈 quantity', '키즈 price',
    '식품 unique', '식품 quantity', '식품 price',
    '스포츠/레저 unique', '스포츠/레저 quantity', '스포츠/레저 price',
    '생활/건강 unique', '생활/건강 quantity', '생활/건강 price',
    '여가/생활편의 unique', '여가/생활편의 quantity', '여가/생활편의 price',
    '리빙 unique', '리빙 quantity', '리빙 price',
    '자동차/공구 unique', '자동차/공구 quantity', '자동차/공구 price',
    '펫 unique', '펫 quantity', '펫 price',
    'e-kam unique', 'e-kam quantity', 'e-kam price',
    '여가/도서(e쿠폰) unique', '여가/도서(e쿠폰) quantity', '여가/도서(e쿠폰) price',
    '노크잇 unique', '노크잇 quantity', '노크잇 price',
]

CATEGORY_COLUMNS  = MEDIA_COLUMNS + CATEGORY_EXTRA_COLUMNS
CATEGORY_MARKER   = '카테고리 구매 unique'  # 카테고리 파일에만 존재하는 컬럼


def detect_file_kind(df):
    """컬럼 구조로 파일 종류 판별. '카테고리 구매 unique' 컬럼 존재 여부가 기준."""
    if df is None:
        return None
    return 'category' if CATEGORY_MARKER in df.columns else 'media'


def _sniff_column_names(uploaded_file, encoding, errors='strict'):
    """헤더 행과 첫 데이터 행의 실제 필드 수를 비교해 컬럼명 리스트를 만든다.
    데이터 행에 트레일링 콤마 등으로 헤더보다 필드가 더 많으면, 그 여분 필드에
    '_extra_N' 자리표시 이름을 붙여준다 — 이렇게 모든 필드에 명시적 이름을 지정해야
    pandas가 이름 없는 첫 열을 인덱스로 흡수하는 것을 막고 Date열이 항상 일반 컬럼으로
    읽히도록 보장할 수 있다."""
    uploaded_file.seek(0)
    header_line = uploaded_file.readline().decode(encoding, errors=errors)
    data_line   = uploaded_file.readline().decode(encoding, errors=errors)

    header_fields = next(csv.reader([header_line])) if header_line else []
    data_fields   = next(csv.reader([data_line]))   if data_line   else header_fields

    n_extra = max(0, len(data_fields) - len(header_fields))
    return header_fields + [f'_extra_{i}' for i in range(n_extra)]


def _read_csv_with_encoding(uploaded_file, encoding, errors='strict'):
    names = _sniff_column_names(uploaded_file, encoding, errors=errors)
    uploaded_file.seek(0)
    df = pd.read_csv(uploaded_file, header=0, names=names, encoding=encoding, encoding_errors=errors)
    extra_cols = [c for c in df.columns if c.startswith('_extra_')]
    if extra_cols:
        df = df.drop(columns=extra_cols)
    return df


def read_csv_robust(uploaded_file):
    """인코딩이 다른 CSV(UTF-8 / CP949 등)를 순차 시도하여 읽는다.
    헤더를 직접 읽어 컬럼명으로 명시 지정해서 읽기 때문에, 데이터 행의 필드 수가
    헤더보다 많은 경우(트레일링 콤마 등)에도 Date 등 첫 열이 인덱스로 잘못 흡수되지
    않고 항상 일반 컬럼으로 읽힌다.

    1) utf-8 / utf-8-sig / cp949 순차 시도
    2) 모두 실패하면 charset-normalizer로 실제 인코딩 자동 감지 후 재시도
    3) 그래도 실패하면 errors='replace'로 깨진 바이트를 대체 문자로 치환해 읽음(경고 표시)
    """
    fname = getattr(uploaded_file, 'name', '(unknown)')
    tried = ('utf-8', 'utf-8-sig', 'cp949')
    for enc in tried:
        try:
            df = _read_csv_with_encoding(uploaded_file, enc)
            st.write(f"[DEBUG] '{fname}' 인코딩 '{enc}'로 읽기 성공")
            return df
        except (UnicodeDecodeError, UnicodeError) as e:
            st.write(f"[DEBUG] '{fname}' 인코딩 '{enc}' 실패: {e}")
            continue

    # 자동 감지: charset-normalizer로 실제 인코딩 추정
    detected = None
    try:
        from charset_normalizer import from_bytes
        uploaded_file.seek(0)
        raw = uploaded_file.read()
        best = from_bytes(raw).best()
        detected = best.encoding if best is not None else None
    except ImportError:
        st.write("[DEBUG] charset-normalizer가 설치되어 있지 않아 자동 감지를 건너뜁니다.")
    except Exception as e:
        st.write(f"[DEBUG] charset-normalizer 감지 중 오류: {e}")

    if detected and detected not in tried:
        st.write(f"[DEBUG] '{fname}' charset-normalizer 감지 인코딩: {detected}")
        try:
            df = _read_csv_with_encoding(uploaded_file, detected)
            st.write(f"[DEBUG] '{fname}' 감지된 인코딩 '{detected}'로 읽기 성공")
            return df
        except (UnicodeDecodeError, UnicodeError, LookupError) as e:
            st.write(f"[DEBUG] '{fname}' 감지된 인코딩 '{detected}'로도 실패: {e}")

    # 최후 수단: 깨진 바이트를 대체 문자로 치환하며 읽기
    st.warning(f"'{fname}'의 인코딩을 확인할 수 없어 일부 문자가 깨진 상태(대체 문자로 치환)로 읽었습니다. 결과를 확인해주세요.")
    return _read_csv_with_encoding(uploaded_file, 'utf-8', errors='replace')


# ──────────────────────────────────────────────────────────────────────────────
# Column-index constants
# ──────────────────────────────────────────────────────────────────────────────

# Media file
M_DATE, M_CAMP, M_ADG, M_AD = ci('A'), ci('F'), ci('G'), ci('H')

# Category file
C_DATE, C_MEDIA, C_CAMP, C_ADG, C_AD = ci('A'), ci('E'), ci('F'), ci('G'), ci('H')
C_AV = ci('AV')                  # 47 – USP Category (사업부-연합 세부 분기용, 원본 그대로 유효)

# 사업부구분(인덱스 그룹 시트 기준) → 카테고리 그룹명. 그룹명 + ' unique'/' quantity'/' price'로
# CATEGORY_EXTRA_COLUMNS의 실제 열 이름을 만든다 (위치가 아닌 이름으로 조회 — BR열 등
# 위치 기반 매핑이 실제 컬럼 구조와 어긋났던 문제 재발 방지).
BIZ_TO_CATEGORY_GROUP = {
    '사업부-가구':          '가구',
    '사업부-그로서리-전체': '식품',
    '사업부-그로서리-별도': '식품',
    '사업부-리빙':          '리빙',
    '사업부-자동차공구':    '자동차/공구',
    '사업부-키즈':          '키즈',
    '사업부-펫':            '펫',
    '사업부-여가생활e쿠폰': '여가/도서(e쿠폰)',
    '사업부-디지털가전':    '디지털가전',
}
USP_TO_CATEGORY_GROUP = {
    'LVG':     '리빙',
    'PET':     '펫',
    'CAR':     '자동차/공구',
    'TOOL':    '자동차/공구',
    'CARTOOL': '자동차/공구',
    'STN':     '자동차/공구',
    'KID':     '키즈',
}
GENERIC_CATEGORY_COLS = ('카테고리 구매 unique', '카테고리 구매 quantity', '카테고리 구매 price')

# '사업부-연합' + USP Category='ALL'인 행이 동일 원본 값을 복사받는 4개 카테고리 그룹.
ALL_UNION_USP_GROUPS = ['리빙', '키즈', '펫', '자동차/공구']


def category_group_cols(group_name):
    return (f'{group_name} unique', f'{group_name} quantity', f'{group_name} price')


# ──────────────────────────────────────────────────────────────────────────────
# 인덱스 파일 (그룹/소재 시트)
# ──────────────────────────────────────────────────────────────────────────────

INDEX_GROUP_COLUMNS    = ['사업부구분', 'Media', 'Campaign', 'Ad Group', '종료일']
INDEX_CREATIVE_COLUMNS = ['사업부구분', 'Media', 'Campaign', 'Ad Group', 'Ad',
                          '프로모션', '기여기간', '종료일']


def parse_index_file(uploaded_file):
    """인덱스 xlsx(그룹/소재/미디어구분 시트)를 읽어 (group_df, creative_df, media_df)로 반환.
    시트명에 '그룹'/'소재'/'미디어'가 포함되면 이를 우선 사용하고, 그룹/소재는 시트명으로
    찾지 못하면 시트 순서(1번째=그룹, 2번째=소재)로 판별한다. [미디어구분] 시트는 없으면
    None으로 둔다(위치 기반 폴백 없음)."""
    sheets = pd.read_excel(uploaded_file, sheet_name=None, header=0, engine='openpyxl')
    group_df = creative_df = media_df = None
    for name, df in sheets.items():
        if '그룹' in str(name):
            group_df = df
        elif '소재' in str(name):
            creative_df = df
        elif '미디어' in str(name):
            media_df = df

    names = list(sheets.keys())
    if group_df is None and len(names) >= 1:
        group_df = sheets[names[0]]
    if creative_df is None and len(names) >= 2:
        creative_df = sheets[names[1]]

    return group_df, creative_df, media_df


INDEX_MISSING_MARK = '#인덱스추가'  # 인덱스에 없는 Campaign+Ad Group(+Ad) 표시값


def build_index_lookup(group_df):
    """그룹 시트 → {(Campaign, Ad Group): (사업부구분, 종료일 Timestamp)} 매핑."""
    lookup = {}
    if group_df is None or group_df.empty:
        return lookup
    if 'Campaign' not in group_df.columns or 'Ad Group' not in group_df.columns:
        return lookup

    biz_col = '사업부구분' if '사업부구분' in group_df.columns else None
    end_col = '종료일'    if '종료일'    in group_df.columns else None

    g = group_df.copy()
    if end_col:
        g[end_col] = parse_flexible_dates(g[end_col])

    for _, row in g.iterrows():
        key = (str(row['Campaign']).strip(), str(row['Ad Group']).strip())
        lookup[key] = (
            row[biz_col] if biz_col else np.nan,
            row[end_col] if end_col else pd.NaT,
        )
    return lookup


def build_creative_lookup(creative_df):
    """소재 시트 → {(Campaign, Ad Group, Ad): 종료일 Timestamp} 매핑."""
    lookup = {}
    if creative_df is None or creative_df.empty:
        return lookup
    if any(c not in creative_df.columns for c in ('Campaign', 'Ad Group', 'Ad')):
        return lookup

    end_col = '종료일' if '종료일' in creative_df.columns else None

    g = creative_df.copy()
    if end_col:
        g[end_col] = parse_flexible_dates(g[end_col])

    for _, row in g.iterrows():
        key = (str(row['Campaign']).strip(), str(row['Ad Group']).strip(), str(row['Ad']).strip())
        lookup[key] = row[end_col] if end_col else pd.NaT
    return lookup


def build_media_group_lookup(media_df):
    """[미디어구분] 시트 → {Media: 대구분} 매핑. 시트가 없거나 비어있으면 빈 딕셔너리를
    반환한다(경고 없음) — [데이터가공]/[D7정제] 탭, 미디어/카테고리 파일 공통으로 사용."""
    lookup = {}
    if media_df is None or media_df.empty:
        return lookup
    if 'Media' not in media_df.columns or '대구분' not in media_df.columns:
        return lookup

    for _, row in media_df.iterrows():
        key = str(row['Media']).strip()
        lookup[key] = row['대구분']
    return lookup


def d7_status(row_date, end_date):
    """종료일+7일 기준 포함/제외 판정. 종료일을 알 수 없으면 기본값 '포함'."""
    if pd.isna(end_date) or pd.isna(row_date):
        return '포함'
    return '포함' if row_date <= end_date + pd.Timedelta(days=7) else '제외'


def add_index_columns(df, group_lookup, creative_lookup, camp_col, adg_col, ad_col, date_col):
    """그룹/소재 인덱스 매칭 결과를 추가한다 — 맨 앞 2열 '그룹_D7초과여부'/'소재_D7초과여부',
    맨 끝 1열 '사업부구분'. 원본 데이터 열 순서는 그대로 유지된다.
    그룹은 Campaign+Ad Group, 소재는 Campaign+Ad Group+Ad로 매칭한다. 행 삭제는 하지 않는다.
    - 그룹 미매칭 → 사업부구분/그룹_D7초과여부 모두 '#인덱스추가'
    - 소재 미매칭이지만 그룹은 매칭 → 소재_D7초과여부 = '그룹기준적용'
    - 그룹·소재 둘 다 미매칭 → 소재_D7초과여부도 '#인덱스추가'
    """
    out   = df.copy()
    dates = parse_flexible_dates(out[date_col])

    biz_list, group_status, creative_status = [], [], []
    for camp, adg, ad, d in zip(out[camp_col], out[adg_col], out[ad_col], dates):
        camp_s, adg_s, ad_s = str(camp).strip(), str(adg).strip(), str(ad).strip()

        gmatch = group_lookup.get((camp_s, adg_s))
        if gmatch is None:
            biz_list.append(INDEX_MISSING_MARK)
            group_status.append(INDEX_MISSING_MARK)
        else:
            biz_val, g_end = gmatch
            biz_list.append(biz_val if pd.notna(biz_val) else INDEX_MISSING_MARK)
            group_status.append(d7_status(d, g_end))

        c_end = creative_lookup.get((camp_s, adg_s, ad_s))
        if c_end is not None:
            creative_status.append(d7_status(d, c_end))
        elif gmatch is not None:
            creative_status.append('그룹기준적용')
        else:
            creative_status.append(INDEX_MISSING_MARK)

    out.insert(0, '소재_D7초과여부', creative_status)
    out.insert(0, '그룹_D7초과여부', group_status)
    out['사업부구분'] = biz_list
    return out


# ──────────────────────────────────────────────────────────────────────────────
# Utilities
# ──────────────────────────────────────────────────────────────────────────────

def safe_col(df, idx):
    return df.columns[idx] if idx < len(df.columns) else None


def row_val(row, idx, default=np.nan):
    return row.iloc[idx] if idx < len(row) else default


DATE_FALLBACK_FORMATS = ('%Y%m%d', '%Y/%m/%d', '%Y.%m.%d', '%Y-%m-%d')


def parse_flexible_dates(series):
    """열에 섞여 있는 여러 날짜 형식(2026-07-03, 20260703, 2026/07/03 등)을 최대한 인식한다.
    pandas.to_datetime은 한 열에 형식이 섞여 있으면 처음 인식한 형식만 유지하고 나머지를
    NaT로 만들어버리는 경우가 있어(형식별로는 개별적으로 모두 정상 파싱됨), 1차 파싱 후
    남은 값에 대해 형식을 하나씩 지정해 재시도한다."""
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    s = series.astype(str).str.strip()
    parsed  = pd.to_datetime(s, errors='coerce')
    missing = parsed.isna() & series.notna()

    for fmt in DATE_FALLBACK_FORMATS:
        if not missing.any():
            break
        attempt = pd.to_datetime(s[missing], format=fmt, errors='coerce')
        parsed.loc[missing] = attempt
        missing = parsed.isna() & series.notna()

    return parsed


# ──────────────────────────────────────────────────────────────────────────────
# Media processing
# ──────────────────────────────────────────────────────────────────────────────

def process_media(raw):
    """업로드된 파일의 모든 데이터를 날짜 범위 필터링 없이 그대로 가공한다."""
    df = raw.copy()
    date_col = df.columns[M_DATE]

    df[date_col] = parse_flexible_dates(df[date_col])
    n_before = len(df)
    df = df.dropna(subset=[date_col]).sort_values(date_col).reset_index(drop=True)

    if df.empty:
        st.warning(f"⚠️ 미디어 파일: A열(날짜)을 인식할 수 있는 행이 없습니다. "
                    f"날짜 형식을 확인해주세요. (원본 {n_before}행 중 0행 인식)")

    return df


# ──────────────────────────────────────────────────────────────────────────────
# Category processing
# ──────────────────────────────────────────────────────────────────────────────

def process_category_step1(raw):
    """1단계: 날짜 파싱/정렬 (컬럼 구조 검증은 호출부에서 처리). 원본 데이터 열은 그대로 유지.
    A열(날짜)을 찾을 수 없거나 인식 가능한 행이 하나도 없으면 ValueError를 낸다."""
    df = raw.copy()
    date_col = safe_col(df, C_DATE)
    if not date_col:
        raise ValueError("A열(날짜)를 찾을 수 없습니다.")

    df[date_col] = parse_flexible_dates(df[date_col])
    df = df.dropna(subset=[date_col]).sort_values(date_col).reset_index(drop=True)

    if df.empty:
        raise ValueError("A열(날짜)을 인식할 수 있는 행이 없습니다. 날짜 형식을 확인해주세요.")

    return df


def process_category_step2(df, group_lookup, creative_lookup):
    """2단계: 인덱스 매칭 — 사업부구분(그룹 시트 Campaign+Ad Group), 그룹_D7초과여부(그룹 시트
    종료일+7일), 소재_D7초과여부(소재 시트 Campaign+Ad Group+Ad)를 계산한다. 매칭되지 않는
    행도 삭제하지 않고 '#인덱스추가'로 표시한다.
    Returns (biz_list, group_status, creative_status) — 모두 len(df) 크기의 object 배열.
    """
    n = len(df)
    biz_list         = np.full(n, '', object)
    group_status     = np.full(n, '', object)
    creative_status  = np.full(n, '', object)

    for pos, (_, row) in enumerate(df.iterrows()):
        camp     = str(row_val(row, C_CAMP, '')).strip()
        adg      = str(row_val(row, C_ADG, '')).strip()
        ad       = str(row_val(row, C_AD, '')).strip()
        row_date = row_val(row, C_DATE)

        gmatch = group_lookup.get((camp, adg))
        if gmatch is None:
            biz_list[pos]     = INDEX_MISSING_MARK
            group_status[pos] = INDEX_MISSING_MARK
        else:
            biz_val, g_end = gmatch
            biz_list[pos]     = biz_val if pd.notna(biz_val) else INDEX_MISSING_MARK
            group_status[pos] = d7_status(row_date, g_end)

        c_end = creative_lookup.get((camp, adg, ad))
        if c_end is not None:
            creative_status[pos] = d7_status(row_date, c_end)
        elif gmatch is not None:
            creative_status[pos] = '그룹기준적용'
        else:
            creative_status[pos] = INDEX_MISSING_MARK

    return biz_list, group_status, creative_status


def process_category_step3(df, biz_list, type_list=None):
    """3단계: 사업부구분(2단계 결과) 기준 카테고리 값 매핑 — 카테고리 구매 unique/quantity/price
    3열의 값만 실제 카테고리 그룹 열(예: '가구 unique/quantity/price')의 값으로 치환하고,
    나머지 45열(패션·뷰티·...·노크잇)은 원본 값 그대로 둔다. 매핑 실패(사업부구분 미매핑,
    사업부-연합의 USP 미매핑, 인덱스 자체 미매칭)는 원본 총계 값을 그대로 두고 수기확인
    플래그만 세운다.

    type_list([데이터가공] 탭 전용, classify_all_rows_v2의 유형 구분 리스트)가 주어지면,
    사업부구분='사업부-그로서리-별도' & 해당 행 유형 구분='카탈로그'인 경우 기본 매핑('식품')
    대신 'e-kam'을 강제 적용한다. type_list=None([D7정제] 탭은 유형 구분 개념이 없어 항상
    None으로 호출됨)이면 이 예외는 적용되지 않고 기존 동작(사업부-그로서리-별도 → 식품)을
    그대로 따른다.

    사업부구분='사업부-연합' & USP Category='ALL'인 행은 단일 그룹이 아니라
    ALL_UNION_USP_GROUPS(리빙/키즈/펫/자동차·공구) 4개 그룹의 unique/quantity/price 열
    전부에 원본 '카테고리 구매' 3열 값을 그대로 복사한다(수기확인 플래그 없음).
    Returns (df, manual_bool_array).
    """
    n      = len(df)
    manual = np.zeros(n, bool)

    generic_u, generic_q, generic_p = GENERIC_CATEGORY_COLS
    # object dtype으로 복사 — pandas 3.x의 엄격한 dtype(예: string[pyarrow])에 다른 타입 값을
    # 대입할 때 발생하는 TypeError를 피하기 위함(앞서 4~7일치 0-처리에서 겪은 문제와 동일한 원인).
    out_u = df[generic_u].astype(object).copy() if generic_u in df.columns else pd.Series([np.nan] * n, dtype=object)
    out_q = df[generic_q].astype(object).copy() if generic_q in df.columns else pd.Series([np.nan] * n, dtype=object)
    out_p = df[generic_p].astype(object).copy() if generic_p in df.columns else pd.Series([np.nan] * n, dtype=object)

    for pos, (_, row) in enumerate(df.iterrows()):
        biz = biz_list[pos]
        biz_str = str(biz).strip() if biz not in ('', None) and biz != INDEX_MISSING_MARK else ''

        group_name = None
        if type_list is not None and biz_str == '사업부-그로서리-별도' \
                and str(type_list[pos]).strip() == '카탈로그':
            group_name = 'e-kam'  # [데이터가공] 탭 전용 예외: 그로서리-별도 + 카탈로그 → e-kam
        elif biz_str == '사업부-연합':
            usp = str(row_val(row, C_AV, '')).strip().upper()
            if usp == 'ALL':
                src_u = row[generic_u] if generic_u in df.columns else np.nan
                src_q = row[generic_q] if generic_q in df.columns else np.nan
                src_p = row[generic_p] if generic_p in df.columns else np.nan
                for union_group in ALL_UNION_USP_GROUPS:
                    gu, gq, gp = category_group_cols(union_group)
                    if gu in df.columns:
                        df.iloc[pos, df.columns.get_loc(gu)] = src_u
                    if gq in df.columns:
                        df.iloc[pos, df.columns.get_loc(gq)] = src_q
                    if gp in df.columns:
                        df.iloc[pos, df.columns.get_loc(gp)] = src_p
                group_name = None  # 카테고리 구매(총계) 열은 원본 값 그대로 유지
            else:
                group_name = USP_TO_CATEGORY_GROUP.get(usp)
                if group_name is None:
                    manual[pos] = True
        elif biz_str:
            group_name = BIZ_TO_CATEGORY_GROUP.get(biz_str)
            if group_name is None:
                manual[pos] = True
        else:
            manual[pos] = True  # 인덱스에 없는 Campaign+Ad Group

        if group_name:
            u_name, q_name, p_name = category_group_cols(group_name)
            out_u.iloc[pos] = row[u_name] if u_name in df.columns else np.nan
            out_q.iloc[pos] = row[q_name] if q_name in df.columns else np.nan
            out_p.iloc[pos] = row[p_name] if p_name in df.columns else np.nan
        # group_name이 없으면(미매칭) 원본 '카테고리 구매' 값(총계 폴백) 그대로 유지

    if generic_u in df.columns:
        df[generic_u] = out_u
        df[generic_q] = out_q
        df[generic_p] = out_p

    return df, manual


def process_category_finalize(df, biz_list, group_status, creative_status):
    """열 순서 재배치 — 소구점 다음에 사업부구분/여백(빈값)/피드구분(빈값) 삽입,
    맨 앞에 그룹_D7초과여부/소재_D7초과여부 삽입. 원본 데이터 열 순서는 그대로 유지."""
    df = df.copy()
    if '소구점' in df.columns:
        pos_after = df.columns.get_loc('소구점') + 1
        df.insert(pos_after, '사업부구분', biz_list)
        df.insert(pos_after + 1, '여백', '')
        df.insert(pos_after + 2, '피드구분', '')
    else:
        df['사업부구분'] = biz_list
        df['여백'] = ''
        df['피드구분'] = ''

    df.insert(0, '소재_D7초과여부', creative_status)
    df.insert(0, '그룹_D7초과여부', group_status)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# 최종 값 치환 — Campaign Theme / 대구분 / USP(ADEF)
#
# [데이터가공]/[D7정제] 탭, 미디어/카테고리 파일 공통으로 각 처리 흐름의 기존 로직이 모두
# 끝난 뒤 엑셀 생성 직전 마지막 단계에서 적용한다. '사업부구분'이 INDEX_MISSING_MARK인
# 행(인덱스 미매칭)은 제외하고 원본 값을 그대로 유지한다.
# ──────────────────────────────────────────────────────────────────────────────

MEDIA_GROUP_MISSING_MARK = '#미디어구분추가필요'  # [미디어구분] 인덱스에 없는 Media 표시값


def apply_final_value_overrides(df, biz_col='사업부구분', media_group_lookup=None):
    """Campaign Theme / Campaign Theme(ADEF) / 대구분 / USP(ADEF) 값을 최종 치환한다.

    - Campaign Theme: 값이 정확히 '-'인 행만 'BS'로 치환
    - Campaign Theme(ADEF): 값이 정확히 '-'인 행만 '사업부'로 치환
    - 대구분: media_group_lookup(Media → 대구분)으로 치환. Media가 lookup에 없으면
      MEDIA_GROUP_MISSING_MARK로 표기
    - USP(ADEF): 값이 정확히 '-'인 행만 '사업부 Static'으로 치환 (USP Category 기반
      LVG/PET/CARTOOL/KID 서브매핑과는 무관한 별도 로직)

    biz_col이 INDEX_MISSING_MARK인 행은 위 3가지 모두에서 제외하고 원본 값을 유지한다.
    Returns (수정된 df, 대구분 미매칭 건수).
    """
    out = df.copy()
    media_group_lookup = media_group_lookup or {}

    if biz_col in out.columns:
        target_mask = out[biz_col] != INDEX_MISSING_MARK
    else:
        target_mask = pd.Series(True, index=out.index)

    if 'Campaign Theme' in out.columns:
        mask = target_mask & (out['Campaign Theme'] == '-')
        out.loc[mask, 'Campaign Theme'] = 'BS'

    if 'Campaign Theme(ADEF)' in out.columns:
        mask = target_mask & (out['Campaign Theme(ADEF)'] == '-')
        out.loc[mask, 'Campaign Theme(ADEF)'] = '사업부'

    n_media_missing = 0
    if '대구분' in out.columns and 'Media' in out.columns:
        new_vals, missing_flags = [], []
        for is_target, media_val, cur_val in zip(target_mask, out['Media'], out['대구분']):
            if not is_target:
                new_vals.append(cur_val)
                missing_flags.append(False)
                continue
            media_key = str(media_val).strip()
            if media_key in media_group_lookup:
                new_vals.append(media_group_lookup[media_key])
                missing_flags.append(False)
            else:
                new_vals.append(MEDIA_GROUP_MISSING_MARK)
                missing_flags.append(True)
        out['대구분'] = new_vals
        n_media_missing = sum(missing_flags)

    if 'USP(ADEF)' in out.columns:
        mask = target_mask & (out['USP(ADEF)'] == '-')
        out.loc[mask, 'USP(ADEF)'] = '사업부 Static'

    return out, n_media_missing


# ──────────────────────────────────────────────────────────────────────────────
# 화면 표시 전 Arrow 직렬화 안전화
# ──────────────────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────────────────────
# Excel export builder
# ──────────────────────────────────────────────────────────────────────────────

def _strip_header_style(writer, sheet_name):
    """헤더(1행) 서식을 일반 텍스트로 초기화 — 색상/볼드 등 어떤 스타일도 없이 출력한다."""
    ws = writer.sheets[sheet_name]
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        cell.font = Font()
        cell.fill = PatternFill()
        cell.border = Border()


def build_media_excel(media_df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        media_df.to_excel(writer, sheet_name='미디어_가공', index=False)
        _strip_header_style(writer, '미디어_가공')
    buf.seek(0)
    return buf.getvalue()


def build_category_excel(cat_df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        cat_df.to_excel(writer, sheet_name='카테고리_전체', index=False)
        _strip_header_style(writer, '카테고리_전체')
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# [데이터가공 탭] 신규 가공 로직 — 캠페인명 단독 매칭 기준
#
# 기존(D7정제 탭) 방식과 완전히 별개의 인덱스 파일(시트: 캠페인 / 피드구분)을 사용한다.
# - [캠페인] 시트: A=Campaign, B=사업부구분, C=유형 구분
# - [피드구분] 시트: A=Creative Full Name, B=피드 구분
#
# RD 파일(미디어/카테고리 공통) 기준 열:
# - F열 = Campaign, AZ열 = Creative Full Name
#
# 그룹_D7초과여부/소재_D7초과여부(A,B열) 계산 자체를 하지 않으므로 별도 삭제 로직이
# 필요 없다 — 애초에 만들지 않는다.
# ──────────────────────────────────────────────────────────────────────────────

TYPE_OTHER_CAMPAIGN = '#그외캠페인'      # 사업부구분 자체가 매칭 안 되는 캠페인
TYPE_NEED_ADD        = '#유형구분추가필요'  # 사업부구분은 매칭되었으나 유형 구분값이 카탈로그/스태틱이 아님
FEED_NEED_ADD        = '#피드구분추가필요'  # 유형 구분=카탈로그인데 피드구분 인덱스가 없음

INDEX2_CAMPAIGN_REQUIRED = ['Campaign', '사업부구분', '유형 구분']
INDEX2_FEED_REQUIRED     = ['Creative Full Name', '피드 구분']

V2_CAMP = ci('F')   # Campaign
V2_CFN  = ci('AZ')  # Creative Full Name


def parse_index_file_v2(uploaded_file):
    """신규 인덱스 xlsx(캠페인/피드구분/미디어구분 시트)를 읽어 (campaign_df, feed_df, media_df)로
    반환. 시트명에 '캠페인'/'피드'/'미디어'가 포함되면 우선 사용하고, 캠페인/피드구분은 시트명으로
    찾지 못하면 시트 순서(1번째=캠페인, 2번째=피드구분)로 판별한다. [미디어구분] 시트는 없으면
    None으로 둔다(위치 기반 폴백 없음)."""
    sheets = pd.read_excel(uploaded_file, sheet_name=None, header=0, engine='openpyxl')
    campaign_df = feed_df = media_df = None
    for name, df in sheets.items():
        if '캠페인' in str(name):
            campaign_df = df
        elif '피드' in str(name):
            feed_df = df
        elif '미디어' in str(name):
            media_df = df

    names = list(sheets.keys())
    if campaign_df is None and len(names) >= 1:
        campaign_df = sheets[names[0]]
    if feed_df is None and len(names) >= 2:
        feed_df = sheets[names[1]]

    return campaign_df, feed_df, media_df


def build_campaign_lookup_v2(campaign_df):
    """[캠페인] 시트 → {Campaign: (사업부구분, 유형 구분)} 매핑."""
    lookup = {}
    if campaign_df is None or campaign_df.empty:
        return lookup
    if 'Campaign' not in campaign_df.columns:
        return lookup

    biz_col  = '사업부구분' if '사업부구분' in campaign_df.columns else None
    type_col = '유형 구분'  if '유형 구분'  in campaign_df.columns else None

    for _, row in campaign_df.iterrows():
        key = str(row['Campaign']).strip()
        biz_val  = row[biz_col]  if biz_col  else np.nan
        type_val = row[type_col] if type_col else np.nan
        lookup[key] = (biz_val, type_val)
    return lookup


def build_feed_lookup_v2(feed_df):
    """[피드구분] 시트 → {Creative Full Name: 피드 구분} 매핑."""
    lookup = {}
    if feed_df is None or feed_df.empty:
        return lookup
    if 'Creative Full Name' not in feed_df.columns:
        return lookup

    feed_col = '피드 구분' if '피드 구분' in feed_df.columns else None

    for _, row in feed_df.iterrows():
        key = str(row['Creative Full Name']).strip()
        lookup[key] = row[feed_col] if feed_col else np.nan
    return lookup


def classify_row_v2(camp, cfn, campaign_lookup, feed_lookup):
    """캠페인명 1건에 대한 (사업부구분, 유형 구분, 피드 구분) 판정.

    - 캠페인명 자체가 인덱스에 없음 → 사업부구분 '#인덱스추가' / 유형·피드 구분 '#그외캠페인'
    - 캠페인명 매칭 O, 유형 구분이 '카탈로그' → 피드구분 시트(Creative Full Name) 매칭
        · 매칭 O → 해당 피드 구분값
        · 매칭 X → '#피드구분추가필요'
    - 캠페인명 매칭 O, 유형 구분이 '스태틱' → 피드 구분 = '스태틱'
    - 캠페인명 매칭 O, 유형 구분이 '카탈로그'/'스태틱' 둘 다 아님(빈값 포함)
        → 유형·피드 구분 모두 '#유형구분추가필요'
    """
    camp_s = str(camp).strip()
    cfn_s  = str(cfn).strip()

    entry = campaign_lookup.get(camp_s)
    if entry is None:
        return INDEX_MISSING_MARK, TYPE_OTHER_CAMPAIGN, TYPE_OTHER_CAMPAIGN

    biz_val, type_val = entry
    biz = str(biz_val).strip() if pd.notna(biz_val) and str(biz_val).strip() else INDEX_MISSING_MARK
    type_str = str(type_val).strip() if pd.notna(type_val) else ''

    if type_str == '카탈로그':
        feed_val = feed_lookup.get(cfn_s)
        if feed_val is None or (isinstance(feed_val, float) and pd.isna(feed_val)) or not str(feed_val).strip():
            feed = FEED_NEED_ADD
        else:
            feed = str(feed_val).strip()
        return biz, '카탈로그', feed
    elif type_str == '스태틱':
        return biz, '스태틱', '스태틱'
    else:
        return biz, TYPE_NEED_ADD, TYPE_NEED_ADD


def classify_all_rows_v2(df, camp_col, cfn_col, campaign_lookup, feed_lookup):
    """DataFrame 전체 행에 대해 (사업부구분 리스트, 유형 구분 리스트, 피드 구분 리스트) 반환."""
    biz_list, type_list, feed_list = [], [], []
    for camp, cfn in zip(df[camp_col], df[cfn_col]):
        biz, typ, feed = classify_row_v2(camp, cfn, campaign_lookup, feed_lookup)
        biz_list.append(biz)
        type_list.append(typ)
        feed_list.append(feed)
    return biz_list, type_list, feed_list


def insert_v2_columns(df, biz_list, type_list, feed_list):
    """소구점 열 바로 다음에 [사업부구분 / 유형 구분 / 피드 구분] 3열을 삽입한다.
    미디어·카테고리 파일 공통으로 사용 — 카테고리 파일은 이 3열 뒤에 기존 48열
    (카테고리 구매~노크잇)이 그대로 이어진다."""
    df = df.copy()
    if '소구점' in df.columns:
        pos = df.columns.get_loc('소구점') + 1
        df.insert(pos, '사업부구분', biz_list)
        df.insert(pos + 1, '유형 구분', type_list)
        df.insert(pos + 2, '피드 구분', feed_list)
    else:
        df['사업부구분'] = biz_list
        df['유형 구분']  = type_list
        df['피드 구분']  = feed_list
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ──────────────────────────────────────────────────────────────────────────────

def _init_session_state():
    defaults = {
        # D7정제 탭(기존)
        'index_group_df':     None,
        'index_creative_df':  None,
        'index_media_df':     None,
        'index_filename':     None,
        'index_uploaded_at':  None,
        'media_log':          None,
        'cat_log':            None,
        # 데이터가공 탭(신규)
        'index2_campaign_df':  None,
        'index2_feed_df':      None,
        'index2_media_df':     None,
        'index2_filename':     None,
        'index2_uploaded_at':  None,
        'media_log2':          None,
        'cat_log2':            None,
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)


# ──────────────────────────────────────────────────────────────────────────────
# 가공 내역(단계별 진행 상태) 패널
# ──────────────────────────────────────────────────────────────────────────────

MEDIA_STEP_LABELS = [
    'CSV 파일 읽기 및 컬럼 구조 검증',
    '인덱스 매칭 — 사업부구분 / D7 판정',
    'Campaign Theme 값 치환',
    '대구분 값 치환 (미디어구분 인덱스)',
    'USP(ADEF) 값 치환',
    '엑셀 파일 생성',
]
CATEGORY_STEP_LABELS = [
    'CSV 파일 읽기 및 컬럼 구조 검증',
    '인덱스 매칭 — 사업부구분 / D7 판정',
    '카테고리 값 매핑 (사업부구분 기준)',
    'Campaign Theme 값 치환',
    '대구분 값 치환 (미디어구분 인덱스)',
    'USP(ADEF) 값 치환',
    '엑셀 파일 생성',
]

MEDIA_STEP_LABELS_V2 = [
    'CSV 파일 읽기 및 컬럼 구조 검증',
    '인덱스 매칭 — 사업부구분 / 유형 구분 / 피드 구분',
    'Campaign Theme 값 치환',
    '대구분 값 치환 (미디어구분 인덱스)',
    'USP(ADEF) 값 치환',
    '엑셀 파일 생성',
]
CATEGORY_STEP_LABELS_V2 = [
    'CSV 파일 읽기 및 컬럼 구조 검증',
    '인덱스 매칭 — 사업부구분 / 유형 구분 / 피드 구분',
    '카테고리 값 매핑 (사업부구분 기준)',
    'Campaign Theme 값 치환',
    '대구분 값 치환 (미디어구분 인덱스)',
    'USP(ADEF) 값 치환',
    '엑셀 파일 생성',
]

# ──────────────────────────────────────────────────────────────────────────────
# 안내 영역 — [데이터가공]/[D7정제] 탭 공통 UI. 탭별 서브탭 이름과 내용만 다르다.
# ──────────────────────────────────────────────────────────────────────────────

GUIDE_CONTENT = {
    'new': [
        ("사업부/유형구분",
         "인덱스 파일 [캠페인] 시트, 캠페인명(F열) 기준 매칭 → 사업부구분 **BR열** / "
         "유형구분 **BS열** 추출\n\n"
         "**사업부구분 추출값**\n"
         "- `각 사업부별 매칭` : 인덱스 파일 기준 매칭\n"
         "- `#인덱스추가` : 인덱스 미매칭 캠페인\n\n"
         "**유형구분 추출값**\n"
         "- `#유형구분추가필요` : 매칭됐지만 값 없음\n"
         "- `#그외캠페인` : 인덱스 미매칭 캠페인"),
        ("피드구분",
         "Creative Full Name(AZ열), [피드구분] 시트 매칭 → 피드구분 **BT열** 추출\n\n"
         "- `각 피드 정상 매칭` : 인덱스 파일 기준 매칭\n"
         "- `스태틱` : 유형구분이 스태틱인 경우 항상 고정\n"
         "- `#피드구분추가필요` : 유형구분이 카탈로그인데 피드 미등록"),
        ("카테고리매핑",
         "- `기본 매핑` : 사업부구분 → 동일 카테고리열 (가구→가구열, 그로서리→식품열 등)\n"
         "- `사업부-연합` : USP Category 값으로 2차 매핑, 미해당 시 `수기확인`\n"
         "- `사업부-연합 + ALL` : USP Category가 `ALL`이면 리빙·키즈·펫·자동차/공구 4개 카테고리 그룹에 동일 값 복사\n"
         "- `그로서리-별도+카탈로그` : 식품열 대신 `e-kam`열로 치환"),
        ("공통치환",
         "- `Campaign Theme` : `-` → `BS`\n"
         "- `Campaign Theme(ADEF)` : `-` → `사업부`\n"
         "- `대구분` : 미디어구분 시트 기준 치환, 미인식 시 `#미디어구분추가필요`\n"
         "- `USP(ADEF)` : `-` → `사업부Static`"),
    ],
    'legacy': [
        ("사업부구분",
         "인덱스 파일 [그룹]/[소재] 시트, Media+Campaign+Ad Group(+Ad) 다중 컬럼 매칭 → "
         "사업부구분 추출\n\n"
         "**추출 위치**\n"
         "- 미디어 파일 : 맨 끝열 (**BT열**)\n"
         "- 카테고리 파일 : 소구점 다음 (**BT열**)\n\n"
         "**추출값**\n"
         "- `각 사업부별 매칭` : 인덱스 파일 기준 매칭\n"
         "- `#인덱스추가` : 인덱스 미매칭 캠페인"),
        ("D7초과여부",
         "데이터 날짜가 인덱스 종료일+7일을 넘는지로 판정 → 미디어·카테고리 공통 맨 앞 2열 추출\n\n"
         "**그룹_D7초과여부 (A열)**\n"
         "- `포함` : 날짜가 [그룹] 시트 종료일+7일 이내\n"
         "- `제외` : 날짜가 [그룹] 시트 종료일+7일 초과\n"
         "- `#인덱스추가` : [그룹] 시트 매칭 없음\n\n"
         "**소재_D7초과여부 (B열)**\n"
         "- `포함` / `제외` : 위와 동일 기준, [소재] 시트 매칭 시\n"
         "- `그룹기준적용` : [소재]는 미매칭, [그룹]은 매칭돼 그룹 기준값 사용\n"
         "- `#인덱스추가` : [그룹] 시트조차 매칭 없음"),
        ("카테고리매핑",
         "- `기본 매핑` : 사업부구분 → 동일 카테고리열\n"
         "- `사업부-연합` : USP Category 값으로 2차 매핑, 미해당 시 `수기확인`\n"
         "- `사업부-연합 + ALL` : USP Category가 `ALL`이면 리빙·키즈·펫·자동차/공구 4개 카테고리 그룹에 동일 값 복사\n"
         "- `그로서리-별도` : e-kam 예외 없음, 식품열 그대로 유지 (데이터가공 탭과의 차이점)"),
        ("공통치환",
         "- `Campaign Theme` : `-` → `BS`\n"
         "- `Campaign Theme(ADEF)` : `-` → `사업부`\n"
         "- `대구분` : 미디어구분 시트 기준 치환, 미인식 시 `#미디어구분추가필요`\n"
         "- `USP(ADEF)` : `-` → `사업부Static`"),
    ],
}


def render_extract_guide(kind):
    """추출값 안내 영역 — "패널"(제목 + 설명) 안에 서브탭(st.tabs) + 탭별 내용을 감싸는
    테두리 박스(st.container(border=True))로 표시한다. 각 탭 내용의 백틱(`) 코드 서식은
    inject_theme_css()에서 레드 톤 칩 스타일로 전역 오버라이드된다.
    kind는 'new'([데이터가공] 탭) 또는 'legacy'([D7정제] 탭)."""
    with st.container(border=True):
        st.markdown("#### 추출값 안내")
        st.caption("가공 과정에서 나오는 추출값이 어떤 의미인지 설명합니다")
        items = GUIDE_CONTENT[kind]
        tabs = st.tabs([name for name, _ in items])
        for tab, (_, content) in zip(tabs, items):
            with tab:
                with st.container(border=True):
                    st.markdown(content)


def inject_theme_css():
    """디자인 레퍼런스(design_concept_v8) 기준 전역 색상/폰트/컴포넌트 스타일 주입.
    브랜드 레드(--primary) 계열 색상, Noto Sans KR 본문 폰트, IBM Plex Mono 데이터성 폰트를
    :root CSS 변수로 정의하고 사이드바 내비게이션·버튼·프로그레스바·단계 트래커·추출값 안내
    칩(코드 서식) 등 세부 요소에 일관 적용한다. .streamlit/config.toml의 theme 설정과 함께
    동작하며, config.toml만으로 커버되지 않는 세부 요소를 여기서 덮어씌운다."""
    st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=Noto+Sans+KR:wght@400;500;600;700&display=swap');

        :root {
            --bg: #F6F8F5;
            --surface: #FFFFFF;
            --primary: #AC332B;
            --primary-dark: #7E241E;
            --mint: #F8E9E7;
            --mint-strong: #F0CFC9;
            --text: #1B2420;
            --text-muted: #6E766F;
            --border: #E3E7E1;
            --radius: 10px;
            --mono: 'IBM Plex Mono', monospace;
            --sans: 'Noto Sans KR', -apple-system, sans-serif;
        }

        html, body, [class*="css"] { font-family: var(--sans) !important; }
        .stApp { background: var(--bg); }
        [data-testid="stAppViewContainer"] { font-family: var(--sans); }

        /* 사이드바 브랜드 블록 */
        .nps-brand { display:flex; align-items:center; gap:10px; padding:2px 4px 18px; }
        .nps-brand-mark {
            width:32px; height:32px; border-radius:8px; background:var(--primary); color:#fff;
            display:flex; align-items:center; justify-content:center;
            font-family:var(--mono); font-weight:500; font-size:13px; flex:0 0 auto;
        }
        .nps-brand-name { font-weight:700; font-size:14.5px; line-height:1.2; color:var(--text); }
        .nps-brand-sub { font-size:11px; color:var(--text-muted); font-family:var(--mono); }
        .nps-nav-group-label {
            font-size:11px; color:var(--text-muted); text-transform:uppercase;
            letter-spacing:.04em; padding:10px 6px 4px; font-weight:600;
        }

        /* 사이드바 내비게이션 버튼 — active(선택됨)는 브랜드 mint 배경 + primary-dark 텍스트 */
        [data-testid="stSidebar"] button {
            text-align:left !important; justify-content:flex-start !important; box-shadow:none !important;
        }
        [data-testid="stSidebar"] button[kind="secondary"] {
            background:transparent !important; border:none !important; color:var(--text) !important; font-weight:500 !important;
        }
        [data-testid="stSidebar"] button[kind="secondary"]:hover { background:var(--bg) !important; }
        [data-testid="stSidebar"] button[kind="primary"] {
            background:var(--mint) !important; color:var(--primary-dark) !important;
            border:none !important; font-weight:600 !important;
        }
        [data-testid="stSidebar"] button[kind="primary"]:hover { background:var(--mint-strong) !important; }

        /* 본문 영역 버튼 — 브랜드 레드 채움(primary) / 테두리만 있는 무채색(secondary) */
        button[kind="primary"], button[data-testid="stBaseButton-primary"] {
            background-color: var(--primary) !important;
            border-color: var(--primary) !important;
            color: #fff !important;
        }
        button[kind="primary"]:hover, button[data-testid="stBaseButton-primary"]:hover {
            background-color: var(--primary-dark) !important;
            border-color: var(--primary-dark) !important;
        }
        button[kind="secondary"], button[data-testid="stBaseButton-secondary"] {
            background-color: var(--surface) !important;
            border-color: var(--border) !important;
            color: var(--text-muted) !important;
        }
        button:disabled, button:disabled:hover {
            background-color: var(--bg) !important;
            border-color: var(--bg) !important;
            color: var(--text-muted) !important;
        }

        [data-testid="stProgressBar"] > div > div { background-color: var(--primary) !important; }

        /* 업로드 카드 */
        .nps-upload-label { font-size:13.5px; font-weight:700; color:var(--text); margin-bottom:8px; }
        .nps-upload-placeholder { font-size:12.5px; color:var(--text-muted); margin:8px 0 10px; }
        .nps-upload-filename {
            font-size:12px; color:var(--text); font-family:var(--mono); margin:8px 0 10px; word-break:break-all;
        }

        /* 진행 현황 패널 — 헤더(제목/진행률/다운로드) */
        .nps-pipeline-title { font-size:14px; font-weight:700; color:var(--text); margin:0; }
        .nps-progress-mini { width:100%; max-width:140px; height:5px; border-radius:3px; background:var(--bg); overflow:hidden; margin-top:6px; }
        .nps-progress-mini > div { height:100%; background:var(--primary); transition:width .2s; }

        /* 가로 스텝 트래커 */
        .nps-track { display:flex; align-items:flex-start; justify-content:space-between; position:relative; gap:4px; margin-top:14px; flex-wrap:wrap; }
        .nps-track::before {
            content:""; position:absolute; top:14px; left:4%; right:4%; height:1px;
            background-image:linear-gradient(to right, var(--border) 50%, transparent 50%);
            background-size:8px 1px; z-index:0;
        }
        .nps-node { display:flex; flex-direction:column; align-items:center; gap:6px; flex:1 1 0; min-width:96px; position:relative; z-index:1; background:var(--surface); padding:0 2px; }
        .nps-node-dot {
            width:28px; height:28px; border-radius:50%; display:flex; align-items:center; justify-content:center;
            font-family:var(--mono); font-size:11px; font-weight:500; border:1.5px solid var(--border);
            background:var(--surface); color:var(--text-muted);
        }
        .nps-node.done .nps-node-dot { background:var(--primary); border-color:var(--primary); color:#fff; }
        .nps-node.active .nps-node-dot { border-color:var(--primary); color:var(--primary-dark); background:var(--mint); }
        .nps-node.error .nps-node-dot { border-color:#B91C1C; color:#fff; background:#B91C1C; }
        .nps-node-label { font-size:11.5px; font-weight:600; text-align:center; color:var(--text); line-height:1.35; }
        .nps-node-sub { font-size:10.5px; color:var(--text-muted); text-align:center; font-family:var(--mono); }
        .nps-node.error .nps-node-sub { color:#B91C1C; }

        /* 추출값 안내 — 백틱 코드 서식을 레드 톤 칩으로 (라인 전체가 아닌 용어에만 배경 적용) */
        [data-testid="stMarkdownContainer"] code {
            font-family: var(--mono) !important;
            font-weight: 500;
            color: var(--primary-dark) !important;
            background: var(--mint) !important;
            padding: 2px 6px;
            border-radius: 5px;
        }
        </style>
    """, unsafe_allow_html=True)


def _new_run_log(step_labels):
    return {
        'steps':       [{'label': s, 'status': 'pending', 'detail': None} for s in step_labels],
        'rows':        None,
        'excel_bytes': None,
        'excel_fname': None,
    }


STEP_STATUS_SUB = {
    'pending': '처리전',
    'running': '처리중',
    'done':    '완료',
    'error':   '오류',
}


def render_upload_card(label, placeholder_text, file_types, uploader_key, button_key, is_index=False):
    """업로드 카드 하나(라벨 + 파일 업로더 + 조건부 버튼)를 그린다.

    - 파일 미업로드 상태: 버튼 = "Upload"(테두리만 있는 무채색 스타일), 비활성화
    - 파일 업로드 후: 버튼 = "가공"(미디어/카테고리) 또는 "완료"(인덱스), 브랜드 레드 채움
    Returns (uploaded_file, button_clicked).
    """
    card_key = f"upload_card_{uploader_key}"
    st.markdown(
        f'<style>.st-key-{card_key} {{ border-style: dashed !important; '
        f'border-color: var(--border) !important; border-radius: var(--radius) !important; }}</style>',
        unsafe_allow_html=True,
    )
    with st.container(key=card_key, border=True):
        st.markdown(f'<div class="nps-upload-label">{escape(label)}</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader(f"{label} 업로드", type=file_types, key=uploader_key,
                                     label_visibility='collapsed')
        if uploaded is not None:
            st.markdown(f'<div class="nps-upload-filename">{escape(uploaded.name)}</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="nps-upload-placeholder">{escape(placeholder_text)}</div>',
                        unsafe_allow_html=True)

        if uploaded is not None:
            btn_label, btn_type = ("완료" if is_index else "가공"), "primary"
        else:
            btn_label, btn_type = "Upload", "secondary"
        clicked = st.button(btn_label, key=button_key, type=btn_type,
                             use_container_width=True, disabled=uploaded is None)
    return uploaded, clicked


def render_step_track_html(steps):
    """단계 진행 상태를 가로 스텝 트래커(노드 연결) HTML로 그린다. 라벨/상태 문자열은 모두
    코드 내부에서 정의된 고정 텍스트이므로 escape 없이 그대로 삽입해도 안전하다(사용자 입력이
    섞이는 detail 텍스트는 여기서 렌더링하지 않고 호출부에서 st.caption으로 별도 표시한다)."""
    state_cls = {'pending': '', 'running': 'active', 'done': 'done', 'error': 'error'}
    nodes = []
    for i, s in enumerate(steps):
        cls = state_cls.get(s['status'], '')
        sub = STEP_STATUS_SUB.get(s['status'], '처리전')
        nodes.append(f'''<div class="nps-node {cls}">
            <div class="nps-node-dot">{i + 1}</div>
            <div class="nps-node-label">{s['label']}</div>
            <div class="nps-node-sub">{sub}</div>
        </div>''')
    return f'<div class="nps-track">{"".join(nodes)}</div>'


def render_run_log(title, log, key_prefix):
    """진행 현황 패널: 제목 + 우측 상단 자체 진행률 바/다운로드 버튼, 그 아래 가로 스텝
    트래커(노드 연결) + 상세 메시지(있는 단계만 caption으로 표시)로 구성한다.

    같은 스크립트 실행 안에서 이 함수가 여러 번 호출되면(단계 진행에 따라 실시간으로
    다시 그릴 때) 컨테이너 key를 고정값으로 두면 Streamlit이 첫 호출 내용에서 갱신을
    멈춰버린다(실측으로 확인된 동작) — 그래서 호출마다 새로운 key를 사용한다.

    이 일련번호는 반드시 st.session_state에 저장해야 한다 — log 딕셔너리에 저장하면
    (예: log['_seq']) 새 가공을 시작할 때마다 _new_run_log()가 완전히 새 log를 만들면서
    번호가 다시 0부터 시작해, 이전 실행 때와 동일한 key가 재사용된다. session_state는
    실행 간에도 유지되므로 여기 저장해야 매 호출이 진짜 유일한 key를 갖는다."""
    steps       = log['steps']
    total       = len(steps)
    done_n      = sum(1 for s in steps if s['status'] == 'done')
    has_error   = any(s['status'] == 'error' for s in steps)
    progress_pct = round(done_n / total * 100) if total else 0
    can_download = bool(log['excel_bytes']) and not has_error

    seq_state_key = f"_{key_prefix}_render_seq"
    seq = st.session_state.get(seq_state_key, 0) + 1
    st.session_state[seq_state_key] = seq

    panel_key = f"{key_prefix}_panel_{seq}"
    dl_key    = f"{key_prefix}_dl_wrap_{seq}"

    st.markdown(f"""
        <style>
        .st-key-{panel_key} {{ border-radius: var(--radius) !important; }}
        .st-key-{dl_key} button {{ padding: 0.25rem 0.9rem; box-shadow: none; }}
        </style>
    """, unsafe_allow_html=True)

    with st.container(key=panel_key, border=True):
        head_l, head_r = st.columns([2.4, 1], vertical_alignment="center")
        with head_l:
            st.markdown(
                f'<div class="nps-pipeline-title">{title}</div>'
                f'<div class="nps-progress-mini"><div style="width:{progress_pct}%;"></div></div>',
                unsafe_allow_html=True,
            )
        with head_r:
            with st.container(key=dl_key):
                st.download_button(
                    "다운로드",
                    data=log['excel_bytes'] or b'',
                    file_name=log['excel_fname'] or 'download.xlsx',
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    disabled=not can_download,
                    use_container_width=True,
                    key=f"{key_prefix}_dl_btn_{seq}",
                )

        st.markdown(render_step_track_html(steps), unsafe_allow_html=True)

        # 상세 메시지(detail)가 있는 단계만 캡션으로 표시 — detail 텍스트는 업로드 파일명 등
        # 사용자 입력을 포함할 수 있어 raw HTML이 아닌 st.caption(자체 escape)으로 렌더링한다.
        for s in steps:
            if s['detail']:
                st.caption(f":red[{s['detail']}]" if s['status'] == 'error' else s['detail'])


def render_new_tab():
    """[데이터가공] 탭 — 캠페인명 단독 매칭 기준 신규 가공 로직.

    - D7 초과 여부 판정 자체를 하지 않음 (그룹_D7초과여부/소재_D7초과여부 열 없음)
    - 인덱스 파일: [캠페인] 시트(Campaign/사업부구분/유형 구분), [피드구분] 시트
      (Creative Full Name/피드 구분) — D7정제 탭의 인덱스와는 완전히 별도로 관리한다.
    """
    u1, u2, u3 = st.columns(3, gap="medium")
    with u1:
        media_file2, run_media2 = render_upload_card(
            "01 · 미디어 파일", "csv 업데이트", ['csv'], 'mf2', 'media_btn2', is_index=False)
    with u2:
        cat_file2, run_cat2 = render_upload_card(
            "02 · 카테고리 파일", "csv 업데이트", ['csv'], 'cf2', 'cat_btn2', is_index=False)
    with u3:
        index_file2, run_index2 = render_upload_card(
            "03 · 인덱스 파일", "xlsx 업데이트", ['xlsx', 'xls'], 'ixf2', 'index_btn2', is_index=True)

    if run_index2:
        if not index_file2:
            st.warning("인덱스 파일을 먼저 선택해주세요.")
        else:
            try:
                campaign_df, feed_df, media_df = parse_index_file_v2(index_file2)
                missing = [c for c in INDEX2_CAMPAIGN_REQUIRED if c not in campaign_df.columns]
                if missing:
                    st.error(f"인덱스 [캠페인] 시트에 필요한 열이 없습니다: {', '.join(missing)}")
                else:
                    if feed_df is None or any(c not in feed_df.columns for c in INDEX2_FEED_REQUIRED):
                        st.warning("인덱스 [피드구분] 시트를 찾지 못했거나 필요한 열이 없습니다. "
                                   "유형 구분이 '카탈로그'인 행은 모두 '#피드구분추가필요'로 표시됩니다.")
                    st.session_state['index2_campaign_df'] = campaign_df
                    st.session_state['index2_feed_df']     = feed_df
                    st.session_state['index2_media_df']    = media_df
                    st.session_state['index2_filename']    = index_file2.name
                    st.session_state['index2_uploaded_at'] = pd.Timestamp.now()
            except Exception as e:
                st.error(f"인덱스 파일 읽기 오류: {e}")

    if st.session_state['index2_campaign_df'] is not None:
        up_at = st.session_state['index2_uploaded_at']
        st.success(f"✅ 현재 적용 중인 인덱스: **{st.session_state['index2_filename']}** "
                   f"({up_at.strftime('%Y-%m-%d %H:%M')})")
        camp_n = len(st.session_state['index2_campaign_df'])
        feed_n = len(st.session_state['index2_feed_df']) \
            if st.session_state['index2_feed_df'] is not None else 0
        st.caption(f"캠페인 시트 {camp_n}건 · 피드구분 시트 {feed_n}건 · 업로드 후에는 다시 올리지 않아도 계속 적용됩니다.")
    else:
        st.info("적용된 인덱스가 없습니다. 파일 선택 후 [완료] 버튼을 눌러주세요.")

    campaign_lookup    = build_campaign_lookup_v2(st.session_state['index2_campaign_df'])
    feed_lookup        = build_feed_lookup_v2(st.session_state['index2_feed_df'])
    media_group_lookup = build_media_group_lookup(st.session_state['index2_media_df'])

    if run_cat2 and cat_file2 and not campaign_lookup:
        st.warning("⚠️ 적용된 인덱스가 없어 사업부구분을 판별할 수 없습니다. "
                   "전체 행이 '#인덱스추가'/'#그외캠페인'으로 표시되고 카테고리 구매 unique/quantity/"
                   "price는 원본 총계 값 그대로 유지됩니다. 인덱스 파일을 먼저 업로드해주세요.")

    media_slot2 = st.empty()
    cat_slot2   = st.empty()

    def _refresh_media2(log, pause=0.15):
        with media_slot2.container():
            render_run_log("미디어 파일 가공", log, key_prefix="media2")
        if pause:
            time.sleep(pause)

    def _refresh_cat2(log, pause=0.15):
        with cat_slot2.container():
            render_run_log("카테고리 파일 가공", log, key_prefix="cat2")
        if pause:
            time.sleep(pause)

    # ────────────────────────────── 미디어 가공 (신규) ──────────────────────────
    if run_media2:
        if not media_file2:
            st.warning("미디어 파일을 업로드해주세요.")
            _refresh_media2(st.session_state['media_log2'] or _new_run_log(MEDIA_STEP_LABELS_V2), pause=0)
        else:
            st.session_state['media_log2'] = None
            log = _new_run_log(MEDIA_STEP_LABELS_V2)
            _refresh_media2(log)

            # 1단계: CSV 파일 읽기 및 컬럼 구조 검증
            m_df = None
            log['steps'][0]['status'] = 'running'
            _refresh_media2(log)
            try:
                media_raw = read_csv_robust(media_file2)
                if detect_file_kind(media_raw) == 'category':
                    raise ValueError(
                        f"'{media_file2.name}'은(는) 카테고리 파일 컬럼 구조로 보입니다. "
                        "카테고리 파일 업로드란에 올린 뒤 카테고리 가공 버튼을 사용해주세요."
                    )
                col_note = None
                if media_raw.shape[1] != len(MEDIA_COLUMNS):
                    col_note = (f"컬럼 수({media_raw.shape[1]})가 예상({len(MEDIA_COLUMNS)}개)과 "
                                "다릅니다. 컬럼 구조를 확인해주세요.")
                m_df = process_media(media_raw)
                if m_df.empty:
                    raise ValueError("A열(날짜)을 인식할 수 있는 행이 없습니다. 날짜 형식을 확인해주세요.")
                log['steps'][0]['status'] = 'done'
                log['steps'][0]['detail'] = col_note
            except Exception as e:
                log['steps'][0]['status'] = 'error'
                log['steps'][0]['detail'] = str(e)
                m_df = None
            _refresh_media2(log)

            # 2단계: 인덱스 매칭 — 사업부구분 / 유형 구분 / 피드 구분
            if log['steps'][0]['status'] == 'done':
                log['steps'][1]['status'] = 'running'
                _refresh_media2(log)
                try:
                    camp_col = m_df.columns[V2_CAMP]
                    cfn_col  = m_df.columns[V2_CFN]
                    biz_list, type_list, feed_list = classify_all_rows_v2(
                        m_df, camp_col, cfn_col, campaign_lookup, feed_lookup
                    )
                    m_df = insert_v2_columns(m_df, biz_list, type_list, feed_list)
                    n_missing = sum(1 for b in biz_list if b == INDEX_MISSING_MARK)
                    log['steps'][1]['status'] = 'done'
                    log['steps'][1]['detail'] = f"#인덱스추가 {n_missing}건" if n_missing else "전체 매칭 완료"
                    log['rows'] = len(m_df)
                except Exception:
                    log['steps'][1]['status'] = 'error'
                    log['steps'][1]['detail'] = traceback.format_exc()
                _refresh_media2(log)

            # 3단계: Campaign Theme 값 치환 / 4단계: 대구분 값 치환 / 5단계: USP(ADEF) 값 치환
            n_media_missing = 0
            if log['steps'][1]['status'] == 'done':
                log['steps'][2]['status'] = 'running'
                _refresh_media2(log)
                try:
                    m_df, n_media_missing = apply_final_value_overrides(
                        m_df, media_group_lookup=media_group_lookup
                    )
                    log['steps'][2]['status'] = 'done'
                except Exception:
                    log['steps'][2]['status'] = 'error'
                    log['steps'][2]['detail'] = traceback.format_exc()
                _refresh_media2(log)

            if log['steps'][2]['status'] == 'done':
                log['steps'][3]['status'] = 'running'
                _refresh_media2(log)
                log['steps'][3]['status'] = 'done'
                log['steps'][3]['detail'] = (
                    f"⚠️ {n_media_missing}건 {MEDIA_GROUP_MISSING_MARK} 발생" if n_media_missing else ""
                )
                _refresh_media2(log)

            if log['steps'][3]['status'] == 'done':
                log['steps'][4]['status'] = 'running'
                _refresh_media2(log)
                log['steps'][4]['status'] = 'done'
                _refresh_media2(log)

            # 6단계: 엑셀 파일 생성
            if log['steps'][4]['status'] == 'done':
                log['steps'][5]['status'] = 'running'
                _refresh_media2(log)
                try:
                    log['excel_bytes'] = build_media_excel(m_df)
                    log['excel_fname'] = f"미디어_가공_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    log['steps'][5]['status'] = 'done'
                except Exception:
                    log['steps'][5]['status'] = 'error'
                    log['steps'][5]['detail'] = traceback.format_exc()
                _refresh_media2(log, pause=0)

            st.session_state['media_log2'] = log
    else:
        _refresh_media2(st.session_state['media_log2'] or _new_run_log(MEDIA_STEP_LABELS_V2), pause=0)

    # ────────────────────────────── 카테고리 가공 (신규) ─────────────────────────
    if run_cat2:
        if not cat_file2:
            st.warning("카테고리 파일을 업로드해주세요.")
            _refresh_cat2(st.session_state['cat_log2'] or _new_run_log(CATEGORY_STEP_LABELS_V2), pause=0)
        else:
            st.session_state['cat_log2'] = None
            log = _new_run_log(CATEGORY_STEP_LABELS_V2)
            _refresh_cat2(log)

            # 1단계: CSV 파일 읽기 및 컬럼 구조 검증
            df1 = None
            log['steps'][0]['status'] = 'running'
            _refresh_cat2(log)
            try:
                cat_raw = read_csv_robust(cat_file2)
                if detect_file_kind(cat_raw) == 'media':
                    raise ValueError(
                        f"'{cat_file2.name}'은(는) 미디어 파일 컬럼 구조로 보입니다. "
                        "미디어 파일 업로드란에 올린 뒤 미디어 가공 버튼을 사용해주세요."
                    )
                col_note = None
                if cat_raw.shape[1] != len(CATEGORY_COLUMNS):
                    col_note = (f"컬럼 수({cat_raw.shape[1]})가 예상({len(CATEGORY_COLUMNS)}개)과 "
                                "다릅니다. 컬럼 구조를 확인해주세요.")
                df1 = process_category_step1(cat_raw)
                log['steps'][0]['status'] = 'done'
                log['steps'][0]['detail'] = col_note
            except Exception as e:
                log['steps'][0]['status'] = 'error'
                log['steps'][0]['detail'] = str(e)
                df1 = None
            _refresh_cat2(log)

            # 2단계: 인덱스 매칭 — 사업부구분 / 유형 구분 / 피드 구분
            biz_list = type_list = feed_list = None
            if log['steps'][0]['status'] == 'done':
                log['steps'][1]['status'] = 'running'
                _refresh_cat2(log)
                try:
                    camp_col = df1.columns[V2_CAMP]
                    cfn_col  = df1.columns[V2_CFN]
                    biz_list, type_list, feed_list = classify_all_rows_v2(
                        df1, camp_col, cfn_col, campaign_lookup, feed_lookup
                    )
                    n_missing = sum(1 for b in biz_list if b == INDEX_MISSING_MARK)
                    log['steps'][1]['status'] = 'done'
                    log['steps'][1]['detail'] = f"#인덱스추가 {n_missing}건" if n_missing else "전체 매칭 완료"
                except Exception:
                    log['steps'][1]['status'] = 'error'
                    log['steps'][1]['detail'] = traceback.format_exc()
                _refresh_cat2(log)

            # 3단계: 카테고리 값 매핑 (사업부구분 기준) — 기존 치환 로직 그대로 재사용
            df2 = None
            if log['steps'][1]['status'] == 'done':
                log['steps'][2]['status'] = 'running'
                _refresh_cat2(log)
                try:
                    biz_array = np.array(biz_list, dtype=object)
                    df2, manual = process_category_step3(df1, biz_array, type_list=type_list)
                    n_manual = int(manual.sum())
                    log['steps'][2]['status'] = 'done'
                    log['steps'][2]['detail'] = f"수기확인 필요 {n_manual}건" if n_manual else "전체 매핑 완료"
                except Exception:
                    log['steps'][2]['status'] = 'error'
                    log['steps'][2]['detail'] = traceback.format_exc()
                _refresh_cat2(log)

            # 4단계: Campaign Theme 값 치환 (여기서 사업부구분/유형구분/피드구분 열을 먼저
            # 삽입해야 apply_final_value_overrides가 사업부구분 기준으로 대상 행을 판정할 수 있음)
            final_df = None
            n_media_missing = 0
            if log['steps'][2]['status'] == 'done':
                log['steps'][3]['status'] = 'running'
                _refresh_cat2(log)
                try:
                    final_df = insert_v2_columns(df2, biz_list, type_list, feed_list)
                    final_df, n_media_missing = apply_final_value_overrides(
                        final_df, media_group_lookup=media_group_lookup
                    )
                    log['rows'] = len(final_df)
                    log['steps'][3]['status'] = 'done'
                except Exception:
                    log['steps'][3]['status'] = 'error'
                    log['steps'][3]['detail'] = traceback.format_exc()
                _refresh_cat2(log)

            # 5단계: 대구분 값 치환
            if log['steps'][3]['status'] == 'done':
                log['steps'][4]['status'] = 'running'
                _refresh_cat2(log)
                log['steps'][4]['status'] = 'done'
                log['steps'][4]['detail'] = (
                    f"⚠️ {n_media_missing}건 {MEDIA_GROUP_MISSING_MARK} 발생" if n_media_missing else ""
                )
                _refresh_cat2(log)

            # 6단계: USP(ADEF) 값 치환
            if log['steps'][4]['status'] == 'done':
                log['steps'][5]['status'] = 'running'
                _refresh_cat2(log)
                log['steps'][5]['status'] = 'done'
                _refresh_cat2(log)

            # 7단계: 엑셀 파일 생성
            if log['steps'][5]['status'] == 'done':
                log['steps'][6]['status'] = 'running'
                _refresh_cat2(log)
                try:
                    log['excel_bytes'] = build_category_excel(final_df)
                    log['excel_fname'] = f"카테고리_가공_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    log['steps'][6]['status'] = 'done'
                except Exception:
                    log['steps'][6]['status'] = 'error'
                    log['steps'][6]['detail'] = traceback.format_exc()
                _refresh_cat2(log, pause=0)

            st.session_state['cat_log2'] = log
    else:
        _refresh_cat2(st.session_state['cat_log2'] or _new_run_log(CATEGORY_STEP_LABELS_V2), pause=0)

    render_extract_guide('new')


def render_legacy_tab():
    """[D7정제] 탭 — 기존 가공 로직은 그대로 보존하고 UI(업로드/안내/가공현황 영역)만 개선."""
    u1, u2, u3 = st.columns(3, gap="medium")
    with u1:
        media_file, run_media = render_upload_card(
            "01 · 미디어 파일", "csv 업데이트", ['csv'], 'mf', 'media_btn', is_index=False)
    with u2:
        cat_file, run_cat = render_upload_card(
            "02 · 카테고리 파일", "csv 업데이트", ['csv'], 'cf', 'cat_btn', is_index=False)
    with u3:
        index_file, run_index = render_upload_card(
            "03 · 인덱스 파일", "xlsx 업데이트", ['xlsx', 'xls'], 'ixf', 'index_btn', is_index=True)

    if run_index:
        if not index_file:
            st.warning("인덱스 파일을 먼저 선택해주세요.")
        else:
            try:
                group_df, creative_df, media_df = parse_index_file(index_file)
                missing = [c for c in INDEX_GROUP_COLUMNS if c not in group_df.columns]
                if missing:
                    st.error(f"인덱스 그룹 시트에 필요한 열이 없습니다: {', '.join(missing)}")
                else:
                    st.session_state['index_group_df']    = group_df
                    st.session_state['index_creative_df'] = creative_df
                    st.session_state['index_media_df']    = media_df
                    st.session_state['index_filename']    = index_file.name
                    st.session_state['index_uploaded_at'] = pd.Timestamp.now()
            except Exception as e:
                st.error(f"인덱스 파일 읽기 오류: {e}")

    if st.session_state['index_group_df'] is not None:
        up_at = st.session_state['index_uploaded_at']
        st.success(f"✅ 현재 적용 중인 인덱스: **{st.session_state['index_filename']}** "
                   f"({up_at.strftime('%Y-%m-%d %H:%M')})")
        g_n = len(st.session_state['index_group_df'])
        c_n = len(st.session_state['index_creative_df']) \
            if st.session_state['index_creative_df'] is not None else 0
        st.caption(f"그룹 시트 {g_n}건 · 소재 시트 {c_n}건 · 업로드 후에는 다시 올리지 않아도 계속 적용됩니다.")
    else:
        st.info("적용된 인덱스가 없습니다. 파일 선택 후 [완료] 버튼을 눌러주세요.")

    group_lookup       = build_index_lookup(st.session_state['index_group_df'])
    creative_lookup    = build_creative_lookup(st.session_state['index_creative_df'])
    media_group_lookup = build_media_group_lookup(st.session_state['index_media_df'])

    # st.empty()로 자리를 잡기 전에 먼저 보여줘야 하는 안내는 여기서 처리한다 — placeholder를
    # 만든 뒤에 일반 st.warning()을 호출하면 placeholder가 이미 차지한 자리보다 아래쪽에
    # 그려져 버려서, 가공 내역 패널보다 밑에 경고가 나오는 순서 뒤바뀜이 생긴다.
    if run_cat and cat_file and not group_lookup:
        st.warning("⚠️ 적용된 인덱스가 없어 사업부구분을 판별할 수 없습니다. "
                   "전체 행이 '#인덱스추가'로 표시되고 카테고리 구매 unique/quantity/price는 "
                   "원본 총계 값 그대로 유지됩니다. 인덱스 파일을 먼저 업로드해주세요.")

    # 가공 내역 영역 — 버튼을 누른 즉시 자리부터 잡아두고(placeholder), 단계가 진행될 때마다
    # 그 자리를 다시 그려서 완료 전에도 실시간으로 진행 상황이 보이도록 한다.
    media_slot = st.empty()
    cat_slot   = st.empty()

    def _refresh_media(log, pause=0.15):
        with media_slot.container():
            render_run_log("미디어 파일 가공", log, key_prefix="media")
        if pause:
            time.sleep(pause)

    def _refresh_cat(log, pause=0.15):
        with cat_slot.container():
            render_run_log("카테고리 파일 가공", log, key_prefix="cat")
        if pause:
            time.sleep(pause)

    # ────────────────────────────── 미디어 가공 ──────────────────────────────
    if run_media:
        if not media_file:
            st.warning("미디어 파일을 업로드해주세요.")
            _refresh_media(st.session_state['media_log'] or _new_run_log(MEDIA_STEP_LABELS), pause=0)
        else:
            st.session_state['media_log'] = None  # 이전 가공(오류 포함) 내역 즉시 초기화
            log = _new_run_log(MEDIA_STEP_LABELS)
            _refresh_media(log)  # 버튼 클릭 즉시 전체 대기 상태로 먼저 표시

            # 1단계: CSV 파일 읽기 및 컬럼 구조 검증
            m_df = None
            log['steps'][0]['status'] = 'running'
            _refresh_media(log)
            try:
                media_raw = read_csv_robust(media_file)
                if detect_file_kind(media_raw) == 'category':
                    raise ValueError(
                        f"'{media_file.name}'은(는) 카테고리 파일 컬럼 구조로 보입니다. "
                        "카테고리 파일 업로드란에 올린 뒤 카테고리 가공 버튼을 사용해주세요."
                    )
                col_note = None
                if media_raw.shape[1] != len(MEDIA_COLUMNS):
                    col_note = (f"컬럼 수({media_raw.shape[1]})가 예상({len(MEDIA_COLUMNS)}개)과 "
                                "다릅니다. 컬럼 구조를 확인해주세요.")
                m_df = process_media(media_raw)
                if m_df.empty:
                    raise ValueError("A열(날짜)을 인식할 수 있는 행이 없습니다. 날짜 형식을 확인해주세요.")
                log['steps'][0]['status'] = 'done'
                log['steps'][0]['detail'] = col_note
            except Exception as e:
                log['steps'][0]['status'] = 'error'
                log['steps'][0]['detail'] = str(e)
                m_df = None
            _refresh_media(log)

            # 2단계: 인덱스 매칭 — 사업부구분 / D7 판정
            if log['steps'][0]['status'] == 'done':
                log['steps'][1]['status'] = 'running'
                _refresh_media(log)
                try:
                    date_col = m_df.columns[M_DATE]
                    camp_col = m_df.columns[M_CAMP]
                    adg_col  = m_df.columns[M_ADG]
                    ad_col   = m_df.columns[M_AD]
                    m_df = add_index_columns(
                        m_df, group_lookup, creative_lookup,
                        camp_col, adg_col, ad_col, date_col
                    )
                    n_missing = int((m_df['사업부구분'] == INDEX_MISSING_MARK).sum())
                    log['steps'][1]['status'] = 'done'
                    log['steps'][1]['detail'] = f"#인덱스추가 {n_missing}건" if n_missing else "전체 매칭 완료"
                    log['rows'] = len(m_df)
                except Exception:
                    log['steps'][1]['status'] = 'error'
                    log['steps'][1]['detail'] = traceback.format_exc()
                _refresh_media(log)

            # 3단계: Campaign Theme 값 치환 / 4단계: 대구분 값 치환 / 5단계: USP(ADEF) 값 치환
            n_media_missing = 0
            if log['steps'][1]['status'] == 'done':
                log['steps'][2]['status'] = 'running'
                _refresh_media(log)
                try:
                    m_df, n_media_missing = apply_final_value_overrides(
                        m_df, media_group_lookup=media_group_lookup
                    )
                    log['steps'][2]['status'] = 'done'
                except Exception:
                    log['steps'][2]['status'] = 'error'
                    log['steps'][2]['detail'] = traceback.format_exc()
                _refresh_media(log)

            if log['steps'][2]['status'] == 'done':
                log['steps'][3]['status'] = 'running'
                _refresh_media(log)
                log['steps'][3]['status'] = 'done'
                log['steps'][3]['detail'] = (
                    f"⚠️ {n_media_missing}건 {MEDIA_GROUP_MISSING_MARK} 발생" if n_media_missing else ""
                )
                _refresh_media(log)

            if log['steps'][3]['status'] == 'done':
                log['steps'][4]['status'] = 'running'
                _refresh_media(log)
                log['steps'][4]['status'] = 'done'
                _refresh_media(log)

            # 6단계: 엑셀 파일 생성
            if log['steps'][4]['status'] == 'done':
                log['steps'][5]['status'] = 'running'
                _refresh_media(log)
                try:
                    log['excel_bytes'] = build_media_excel(m_df)
                    log['excel_fname'] = f"미디어_가공_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    log['steps'][5]['status'] = 'done'
                except Exception:
                    log['steps'][5]['status'] = 'error'
                    log['steps'][5]['detail'] = traceback.format_exc()
                _refresh_media(log, pause=0)

            st.session_state['media_log'] = log
    else:
        _refresh_media(st.session_state['media_log'] or _new_run_log(MEDIA_STEP_LABELS), pause=0)

    # ────────────────────────────── 카테고리 가공 ────────────────────────────
    if run_cat:
        if not cat_file:
            st.warning("카테고리 파일을 업로드해주세요.")
            _refresh_cat(st.session_state['cat_log'] or _new_run_log(CATEGORY_STEP_LABELS), pause=0)
        else:
            st.session_state['cat_log'] = None  # 이전 가공(오류 포함) 내역 즉시 초기화
            log = _new_run_log(CATEGORY_STEP_LABELS)
            _refresh_cat(log)  # 버튼 클릭 즉시 전체 대기 상태로 먼저 표시

            # 1단계: CSV 파일 읽기 및 컬럼 구조 검증
            df1 = None
            log['steps'][0]['status'] = 'running'
            _refresh_cat(log)
            try:
                cat_raw = read_csv_robust(cat_file)
                if detect_file_kind(cat_raw) == 'media':
                    raise ValueError(
                        f"'{cat_file.name}'은(는) 미디어 파일 컬럼 구조로 보입니다. "
                        "미디어 파일 업로드란에 올린 뒤 미디어 가공 버튼을 사용해주세요."
                    )
                col_note = None
                if cat_raw.shape[1] != len(CATEGORY_COLUMNS):
                    col_note = (f"컬럼 수({cat_raw.shape[1]})가 예상({len(CATEGORY_COLUMNS)}개)과 "
                                "다릅니다. 컬럼 구조를 확인해주세요.")
                df1 = process_category_step1(cat_raw)
                log['steps'][0]['status'] = 'done'
                log['steps'][0]['detail'] = col_note
            except Exception as e:
                log['steps'][0]['status'] = 'error'
                log['steps'][0]['detail'] = str(e)
                df1 = None
            _refresh_cat(log)

            # 2단계: 인덱스 매칭 — 사업부구분 / D7 판정
            biz_list = group_status = creative_status = None
            if log['steps'][0]['status'] == 'done':
                log['steps'][1]['status'] = 'running'
                _refresh_cat(log)
                try:
                    biz_list, group_status, creative_status = process_category_step2(
                        df1, group_lookup, creative_lookup
                    )
                    n_missing = int((biz_list == INDEX_MISSING_MARK).sum())
                    log['steps'][1]['status'] = 'done'
                    log['steps'][1]['detail'] = f"#인덱스추가 {n_missing}건" if n_missing else "전체 매칭 완료"
                except Exception:
                    log['steps'][1]['status'] = 'error'
                    log['steps'][1]['detail'] = traceback.format_exc()
                _refresh_cat(log)

            # 3단계: 카테고리 값 매핑 (사업부구분 기준)
            df2 = None
            if log['steps'][1]['status'] == 'done':
                log['steps'][2]['status'] = 'running'
                _refresh_cat(log)
                try:
                    df2, manual = process_category_step3(df1, biz_list)
                    n_manual = int(manual.sum())
                    log['steps'][2]['status'] = 'done'
                    log['steps'][2]['detail'] = f"수기확인 필요 {n_manual}건" if n_manual else "전체 매핑 완료"
                except Exception:
                    log['steps'][2]['status'] = 'error'
                    log['steps'][2]['detail'] = traceback.format_exc()
                _refresh_cat(log)

            # 4단계: Campaign Theme 값 치환 (여기서 최종 열 배치를 먼저 끝내야
            # apply_final_value_overrides가 사업부구분 기준으로 대상 행을 판정할 수 있음)
            final_df = None
            n_media_missing = 0
            if log['steps'][2]['status'] == 'done':
                log['steps'][3]['status'] = 'running'
                _refresh_cat(log)
                try:
                    final_df = process_category_finalize(df2, biz_list, group_status, creative_status)
                    final_df, n_media_missing = apply_final_value_overrides(
                        final_df, media_group_lookup=media_group_lookup
                    )
                    log['rows'] = len(final_df)
                    log['steps'][3]['status'] = 'done'
                except Exception:
                    log['steps'][3]['status'] = 'error'
                    log['steps'][3]['detail'] = traceback.format_exc()
                _refresh_cat(log)

            # 5단계: 대구분 값 치환
            if log['steps'][3]['status'] == 'done':
                log['steps'][4]['status'] = 'running'
                _refresh_cat(log)
                log['steps'][4]['status'] = 'done'
                log['steps'][4]['detail'] = (
                    f"⚠️ {n_media_missing}건 {MEDIA_GROUP_MISSING_MARK} 발생" if n_media_missing else ""
                )
                _refresh_cat(log)

            # 6단계: USP(ADEF) 값 치환
            if log['steps'][4]['status'] == 'done':
                log['steps'][5]['status'] = 'running'
                _refresh_cat(log)
                log['steps'][5]['status'] = 'done'
                _refresh_cat(log)

            # 7단계: 엑셀 파일 생성
            if log['steps'][5]['status'] == 'done':
                log['steps'][6]['status'] = 'running'
                _refresh_cat(log)
                try:
                    log['excel_bytes'] = build_category_excel(final_df)
                    log['excel_fname'] = f"카테고리_가공_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    log['steps'][6]['status'] = 'done'
                except Exception:
                    log['steps'][6]['status'] = 'error'
                    log['steps'][6]['detail'] = traceback.format_exc()
                _refresh_cat(log, pause=0)

            st.session_state['cat_log'] = log
    else:
        _refresh_cat(st.session_state['cat_log'] or _new_run_log(CATEGORY_STEP_LABELS), pause=0)

    render_extract_guide('legacy')


NAV_PAGES = {
    'new': {
        'nav_label': '데이터가공',
        'title':     '데이터가공',
        'sub':       '캠페인명(F열) 단일 매칭 기준으로 사업부구분·유형구분·피드구분을 분류합니다',
        'render':    render_new_tab,
    },
    'legacy': {
        'nav_label': 'D7정제',
        'title':     'D7정제',
        'sub':       'Media+Campaign+Ad Group 다중 매칭 기준으로 사업부구분·D7 초과 여부를 판별합니다',
        'render':    render_legacy_tab,
    },
}


def render_sidebar_nav():
    """사이드바 상단 브랜드("데이터 가공기") + 내비게이션(데이터가공/D7정제). 선택된 메뉴는
    st.session_state['nav_page']에 저장하고 브랜드 mint 배경으로 강조한다."""
    st.session_state.setdefault('nav_page', 'new')

    with st.sidebar:
        st.markdown(
            '<div class="nps-brand">'
            '<div class="nps-brand-mark">NPS</div>'
            '<div><div class="nps-brand-name">데이터 가공기</div>'
            '<div class="nps-brand-sub">wisebirds</div></div>'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown('<div class="nps-nav-group-label">작업</div>', unsafe_allow_html=True)
        for key, meta in NAV_PAGES.items():
            active = st.session_state['nav_page'] == key
            if st.button(meta['nav_label'], key=f'nav_{key}',
                         type="primary" if active else "secondary",
                         use_container_width=True):
                st.session_state['nav_page'] = key
                st.rerun()

    return st.session_state['nav_page']


def main():
    st.set_page_config(page_title="NPS Report 가공기", layout="wide", page_icon="📊")

    inject_theme_css()
    _init_session_state()

    page = render_sidebar_nav()
    meta = NAV_PAGES[page]

    st.markdown(
        f'<div style="margin-bottom:20px;">'
        f'<h1 style="font-size:20px;font-weight:700;margin:0 0 4px;">{meta["title"]}</h1>'
        f'<p style="margin:0;font-size:13px;color:var(--text-muted);">{meta["sub"]}</p>'
        f'</div>',
        unsafe_allow_html=True,
    )

    meta['render']()


if __name__ == "__main__":
    main()
